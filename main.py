import os
import json
import shutil
import uuid
from typing import Optional, Dict
import secrets
from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException, BackgroundTasks, Depends, status
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import uvicorn
from dotenv import load_dotenv
import openpyxl
import google.generativeai as old_genai # Keeping old for safety if needed, but primary is new
from google import genai
from google.genai import types
from openpyxl.styles import Alignment
import datetime

load_dotenv()

# Configure Gemini (New SDK)
GENAI_API_KEY = os.getenv("GEMINI_API_KEY")
client = None
if GENAI_API_KEY:
    try:
        client = genai.Client(api_key=GENAI_API_KEY)
    except Exception as e:
        print(f"Failed to initialize GenAI Client: {e}")

# Security Configuration
APP_USERNAME = os.getenv("APP_USERNAME", "admin")
APP_PASSWORD = os.getenv("APP_PASSWORD", "password")
security = HTTPBasic()

def get_current_username(credentials: HTTPBasicCredentials = Depends(security)):
    current_username_bytes = credentials.username.encode("utf8")
    correct_username_bytes = APP_USERNAME.encode("utf8")
    is_correct_username = secrets.compare_digest(
        current_username_bytes, correct_username_bytes
    )
    
    current_password_bytes = credentials.password.encode("utf8")
    correct_password_bytes = APP_PASSWORD.encode("utf8")
    is_correct_password = secrets.compare_digest(
        current_password_bytes, correct_password_bytes
    )
    
    if not (is_correct_username and is_correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username

# File Cleanup Helper
def cleanup_files(file_paths: list[str]):
    """Delete files from the filesystem."""
    for path in file_paths:
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"Deleted temp file: {path}")
        except Exception as e:
            print(f"Error deleting file {path}: {e}")

app = FastAPI()

# Setup templates
templates = Jinja2Templates(directory="templates")

# Mount static files (for favicon, manifest, etc.)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Directories
TEMP_DIR = "temp"
OUTPUT_DIR = "outputs"
STATIC_DIR = "static"
os.makedirs(TEMP_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(STATIC_DIR, exist_ok=True)

# Load Configuration
with open("mapping_config.json", "r", encoding="utf-8") as f:
    TEMPLATE_CONFIG = json.load(f)

# Helper: Read Interim Monitoring Data
def read_excel_monitoring_data(excel_path: str) -> str:
    """Read interim monitoring data from an uploaded Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # Try to find the interim sheet
        sheet_name = "モニタリング(中間)"
        if sheet_name not in wb.sheetnames:
            # Fallback to first sheet if specific name not found
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        
        # Extract key data based on known cell positions
        data_lines = []
        data_lines.append(f"利用者氏名: {ws['C3'].value or ''}")
        data_lines.append(f"利用者氏名_ふりがな: {ws['C2'].value or ''}")
        data_lines.append(f"作成者: {ws['J2'].value or ''}")
        data_lines.append(f"作成年月日: {ws['K3'].value or ''}{ws['L3'].value or ''}{ws['M3'].value or ''}")
        
        # Goal 1
        data_lines.append(f"達成目標: {ws['B6'].value or ''}")
        status1 = ws['E6'].value or ws['F6'].value or ws['G6'].value or "未定"
        data_lines.append(f"達成状況: {status1}")
        data_lines.append(f"未達成原因・分析1: {ws['H6'].value or ''}")
        data_lines.append(f"今後の対応: {ws['L6'].value or ''}")
        
        # Goal 2
        data_lines.append(f"達成目標: {ws['B11'].value or ''}")
        status2 = ws['E11'].value or ws['F11'].value or ws['G11'].value or "未定"
        data_lines.append(f"達成状況: {status2}")
        data_lines.append(f"未達成原因・分析2: {ws['H11'].value or ''}")
        data_lines.append(f"今後の対応: {ws['L11'].value or ''}")
        
        # Goal 3
        data_lines.append(f"達成目標: {ws['B16'].value or ''}")
        status3 = ws['E16'].value or ws['F16'].value or ws['G16'].value or "未定"
        data_lines.append(f"達成状況: {status3}")
        data_lines.append(f"未達成原因・分析3: {ws['H16'].value or ''}")
        data_lines.append(f"今後の対応: {ws['L16'].value or ''}")
        
        # Other notes
        data_lines.append(f"その他の気づき: {ws['A22'].value or ''}")
        
        return "\n".join(data_lines)
    except Exception as e:
        print(f"Error reading interim Excel: {e}")
        return ""

# Helper: Fill Excel
def fill_excel(template_path: str, mapping: Dict[str, str], config_mapping: Dict[str, str], output_name: str = None) -> str:
    """Fill the Excel template with data based on config mapping."""
    wb = openpyxl.load_workbook(template_path)
    
    # Select sheet (Default)
    default_sheet_name = mapping.pop("_sheet_name", None)
    if default_sheet_name and default_sheet_name in wb.sheetnames:
         default_sheet = wb[default_sheet_name]
    else:
         default_sheet = wb.active
    
    # Invert the config mapping to know which AI key goes to which Cell
    # config_mapping is { "Label": "Cell" }
    # AI returns { "Label": "Value" }
    
    for label, value in mapping.items():
        if label in config_mapping:
            config_value = config_mapping[label]
            
            # Determine target sheet and cell
            target_sheet = default_sheet
            cell_coord = config_value
            
            # Check if config_value has "SheetName!Cell" format
            if "!" in config_value:
                parts = config_value.split("!")
                if len(parts) == 2:
                    s_name, c_coord = parts
                    if s_name in wb.sheetnames:
                        target_sheet = wb[s_name]
                        cell_coord = c_coord
            
            try:
                # Skip if value is None (preserves template content)
                if value is None:
                    continue

                target_sheet[cell_coord] = value
                
                # Check for Vertical Text requirement (Status fields starting with '【')
                # If value starts with '【' and is short (e.g. "【達成】"), assume vertical alignment needed.
                if isinstance(value, str) and value.startswith("【") and len(value) < 10:
                    current_align = target_sheet[cell_coord].alignment
                    new_align = Alignment(
                        horizontal='center', # Center alignment looks best for vertical
                        vertical='center',
                        text_rotation=255,   # 255 = Vertical Text (Stacked)
                        wrap_text=True,      # Often good to keep on standard vertical cells
                        shrink_to_fit=current_align.shrink_to_fit,
                        indent=current_align.indent
                    )
                    target_sheet[cell_coord].alignment = new_align
                
                # Enable text wrapping for long content cells
                elif isinstance(value, str) and len(value) > 50:
                    current_align = target_sheet[cell_coord].alignment
                    new_align = Alignment(
                        horizontal=current_align.horizontal or 'left',
                        vertical=current_align.vertical or 'top',
                        wrap_text=True,  # Enable text wrapping
                        shrink_to_fit=False,  # Disable shrink to fit
                        indent=current_align.indent
                    )
                    target_sheet[cell_coord].alignment = new_align
            except Exception as e:
                print(f"Error writing to {cell_coord} ({label}): {e}")
            
    if output_name:
        output_filename = output_name
    else:
        output_filename = f"processed_{uuid.uuid4().hex}.xlsx"
        
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    wb.save(output_path)
    return output_filename

# Helper: Gemini Processing
def call_gemini(template_info: dict, text_input: str = None, file_paths: list = [], interim_data: str = None) -> Dict[str, str]:
    """
    Call Gemini to map input data to Excel structure.
    interim_data: Optional string containing interim monitoring data for final evaluation mode.
    """
    if not client:
        raise HTTPException(status_code=500, detail="Gemini Client not initialized.")

    # Construct mappings description for the prompt - ONLY field names, NO cell addresses
    mapping_keys = "\n".join([f"- {key}" for key in template_info['mapping'].keys()])

    # Get context (document purpose/meaning) if available
    context_instruction = template_info.get('context', "")
    if context_instruction:
        context_instruction = f"\n\n--- Document Context ---\n{context_instruction}\n------------------------\n"

    # Get style instruction if available
    style_instruction = template_info.get('style_instruction', "")
    if style_instruction:
        style_instruction = f"\n\n--- Writing Style & Formatting Rules ---\n{style_instruction}\n----------------------------------------\n"

    # Construct the prompt text
    system_instruction = (
        "You are an expert welfare record assistant specializing in Japanese disability welfare services (障害福祉サービス). "
        "Your task is to understand the provided audio/images/text and extract relevant information for official documentation.\n"
        f"{context_instruction}\n"
        "You will receive:\n"
        "1. A list of target fields to extract.\n"
        "2. Input data (Audio, PDF, Images, or Text).\n\n"
        "Instructions:\n"
        "- Thoroughly analyze ALL input data to understand the context and meaning.\n"
        "- Extract information that semantically matches each target field, even if exact wording differs.\n"
        "- Map the extracted information to the following target fields:\n"
        f"{mapping_keys}\n"
        f"{style_instruction}\n"
        "- Return ONLY a valid JSON object where keys are the EXACT Field Names provided above and values are the extracted content.\n"
        "- IMPORTANT: Use the field names exactly as listed above as your JSON keys. Do NOT use any other format.\n"
        "- If a key contains '_チェック' (underscore check), output the string '✓' if the condition is true/present, otherwise leave it empty.\n"
        "- If a piece of information is missing, leave the value as an empty string or null.\n"
        "- Do not include markdown formatting (like ```json), just the raw JSON string.\n"
    )
    
    contents = [system_instruction]
    
    # Add interim monitoring data if provided (for final evaluation mode)
    if interim_data:
        contents.append(f"--- 中間評価時のデータ (Interim Monitoring Data) ---\n{interim_data}\n--- 中間評価データここまで ---\n")
    
    if text_input:
        contents.append(f"--- Input Data (Text) ---\n{text_input}\n")
    
    for path in file_paths:
        mime_type = "application/octet-stream" # Default
        if path.lower().endswith(".mp3"): mime_type = "audio/mp3"
        elif path.lower().endswith(".wav"): mime_type = "audio/wav"
        elif path.lower().endswith(".m4a"): mime_type = "audio/mp4"
        elif path.lower().endswith(".aac"): mime_type = "audio/aac"
        elif path.lower().endswith(".flac"): mime_type = "audio/flac"
        elif path.lower().endswith(".ogg"): mime_type = "audio/ogg"
        elif path.lower().endswith(".jobt"): mime_type = "image/jpeg" 
        elif path.lower().endswith(".jpg") or path.lower().endswith(".jpeg"): mime_type = "image/jpeg"
        elif path.lower().endswith(".png"): mime_type = "image/png"
        elif path.lower().endswith(".pdf"): mime_type = "application/pdf"
        elif path.lower().endswith(".txt"): mime_type = "text/plain"

        # Read file bytes for new SDK upload
        print(f"Reading file: {path}")
        with open(path, "rb") as f:
            file_data = f.read()
            
        part = types.Part.from_bytes(data=file_data, mime_type=mime_type)
        contents.append(part)
        
    contents.append("\nExtract the information and map it to the JSON structure.")

    print("Sending request to Gemini (v3 Flash Preview)...")
    try:
        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=contents
        )
    except Exception as e:
        print(f"Gemini API Error: {e}")
        raise HTTPException(status_code=500, detail=f"Gemini API Error: {str(e)}")
    
    # Parse JSON
    try:
        cleaned_text = response.text.strip()
        if cleaned_text.startswith("```json"):
            cleaned_text = cleaned_text[7:]
        if cleaned_text.endswith("```"):
            cleaned_text = cleaned_text[:-3]
        
        # DEBUG: Save response to file
        with open("debug_last_response.json", "w", encoding="utf-8") as f:
            f.write(cleaned_text)
            
        mapping = json.loads(cleaned_text)
        return mapping
    except Exception as e:
        print(f"Error parsing Gemini response: {response.text}")
        raise HTTPException(status_code=500, detail="Failed to interpret AI response.")

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request, username: str = Depends(get_current_username)):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/templates")
async def get_templates(username: str = Depends(get_current_username)):
    """Return available templates."""
    return TEMPLATE_CONFIG

@app.post("/process")
async def process_data(
    background_tasks: BackgroundTasks,
    template_id: str = Form(...),
    text_input: str = Form(None),
    user_name: str = Form(None),
    user_name_furigana: str = Form(None),
    staff_name: str = Form(None),
    date: str = Form(None),
    location: str = Form(None),
    time: str = Form(None),
    count: str = Form(None),
    next_date: str = Form(None),
    cm_location: str = Form(None),
    cm_time: str = Form(None),
    cm_attendees: str = Form(None),
    cm_service_manager: str = Form(None),
    files: list[UploadFile] = File(None),
    username: str = Depends(get_current_username)
):
    if template_id not in TEMPLATE_CONFIG:
        raise HTTPException(status_code=400, detail="Invalid template ID")
    
    selected_template = TEMPLATE_CONFIG[template_id]
    
    # ... template path logic ... (skip lines 147-160, assuming they are unchanged for now, I need to match context)
    # Actually I should be careful not to delete existing logic. 
    # I will construct the meta_info string first.
    
    # Construct Manual Info String
    manual_info_list = []
    if user_name: manual_info_list.append(f"利用者名 (User Name): {user_name}")
    if staff_name: manual_info_list.append(f"作成担当者 (Staff Name): {staff_name}")
    if date: manual_info_list.append(f"日付 (Date): {date}")
    if location: manual_info_list.append(f"開催場所 (Location): {location}")
    if time: manual_info_list.append(f"時間 (Time): {time}")
    if count: manual_info_list.append(f"回数 (Count): {count}")
    if next_date: manual_info_list.append(f"次回予定 (Next Date): {next_date}")
    # Casemeeting-specific
    if cm_service_manager: manual_info_list.append(f"サービス管理責任者 (Service Manager): {cm_service_manager}")
    if cm_location: manual_info_list.append(f"開催場所 (Location): {cm_location}")
    if cm_time: manual_info_list.append(f"開催時間 (Time): {cm_time}")
    if cm_attendees: manual_info_list.append(f"会議出席者 (Attendees): {cm_attendees}")
    
    manual_info_text = ""
    if manual_info_list:
        manual_info_text = "\n\n【基本情報 (Basic Information provided by User)】\n" + "\n".join(manual_info_list) + "\n"
        # Prioritize manual info
        manual_info_text += "IMPORTANT: Please use the above 'Basic Information' to fill the corresponding fields in the output JSON.\n"

    # Combine with text_input
    full_text_input = (text_input or "") + manual_info_text
    
    # 1. Load Excel Template
    # Template files are expected to be in 'welfare_record_app/template/' 
    # but the config says "template/filename.xlsx", so it depends on CWD.
    # Assuming CWD is 'welfare_record_app'.
    template_source_path = selected_template['filename']
    if not os.path.exists(template_source_path):
         # Try prepending 'template/' or checking relative path
         # Just incase config just has filename
         alt_path = os.path.join("template", os.path.basename(selected_template['filename']))
         if os.path.exists(alt_path):
             template_source_path = alt_path
         else:
             raise HTTPException(status_code=500, detail=f"Template file not found: {template_source_path}")

    # 2. Handle Input Data & Call Gemini
    file_paths = []
    
    try:
        # Save uploaded files
        if files:
            for file in files:
                # Skip empty filenames
                if not file.filename: continue
                
                file_path = os.path.join(TEMP_DIR, f"{uuid.uuid4()}_{file.filename}")
                with open(file_path, "wb") as f:
                    shutil.copyfileobj(file.file, f)
                file_paths.append(file_path)
        
        if not full_text_input and not file_paths:
             raise HTTPException(status_code=400, detail="No input provided (files or text).")

        # Special handling for monitoring_final: extract interim data from uploaded Excel
        interim_data = None
        if template_id == "monitoring_final":
            for fp in file_paths:
                # Check if it's an Excel file
                if fp.endswith('.xlsx') or fp.endswith('.xls'):
                    interim_data = read_excel_monitoring_data(fp)
                    if interim_data:
                        break  # Use first Excel file found
        
        mapping = call_gemini(selected_template, text_input=full_text_input, file_paths=file_paths, interim_data=interim_data)
        
        # --- PRIORITY OVERRIDE START ---
        # Overwrite AI results with Manual Inputs if provided
        if user_name and user_name.strip():
            mapping["利用者氏名"] = user_name
            mapping["氏名"] = user_name # Fallback key
            
        if user_name_furigana and user_name_furigana.strip():
            mapping["利用者氏名_ふりがな"] = user_name_furigana
            mapping["氏名のふりがな"] = user_name_furigana

        if staff_name and staff_name.strip():
            mapping["作成者"] = staff_name
            
        if date and date.strip():
            # Input date is YYYY-MM-DD (e.g. 2026-05-20)
            try:
                dt = datetime.datetime.strptime(date, "%Y-%m-%d")
                mapping["作成年_西暦"] = f"{dt.year}年"
                mapping["作成月"] = f"{dt.month}月"
                mapping["作成日"] = f"{dt.day}日"
                # Fallbacks for other templates
                mapping["作成年月日"] = dt.strftime("%Y年%m月%d日")
                mapping["日付"] = dt.strftime("%Y年%m月%d日")
                mapping["実施日"] = dt.strftime("%Y年%m月%d日")
                # Casemeeting date format
                mapping["開催日（令和〇年〇月〇日）"] = f"令和{dt.year - 2018}年{dt.month}月{dt.day}日"
            except ValueError:
                pass
                
        # Casemeeting-specific priority overrides
        if cm_location and cm_location.strip():
            mapping["開催場所"] = cm_location
            
        if cm_time and cm_time.strip():
            mapping["開催時間"] = cm_time
            
        if cm_attendees and cm_attendees.strip():
            mapping["会議出席者"] = cm_attendees
            
        # Casemeeting uses 利用者様 (with 様)
        if user_name and user_name.strip():
            mapping["利用者様"] = user_name
            
        if cm_service_manager and cm_service_manager.strip():
            mapping["サービス管理責任者"] = cm_service_manager
        # --- PRIORITY OVERRIDE END ---
            
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail=f"AI Processing failed: {str(e)}")
    
    # 3. Fill Excel
    # 3. Fill Excel
    try:
        if mapping:
            mapping['_sheet_name'] = selected_template.get('sheet_name')
        
        # Determine User Name for filename
        user_name_val = "名称未設定"
        possible_keys = ["氏名", "利用者名", "利用者様", "利用者氏名"]
        for key in possible_keys:
            if key in mapping and mapping[key]:
                user_name_val = mapping[key]
                break
        if user_name_val == "名称未設定" and "氏名のふりがな" in mapping:
             user_name_val = mapping["氏名のふりがな"]

        # Date string
        date_str = datetime.datetime.now().strftime("%y.%m.%d")
        template_name = selected_template['name']
        
        # Construct Filename: YY.MM.DD_TemplateName【UserName】.xlsx
        # Ensure safe filename
        safe_user_name = "".join([c for c in user_name_val if c.isalnum() or c in (' ', '　', '_', '-')])
        custom_filename = f"{date_str}_{template_name}【{safe_user_name}】.xlsx"
        
        output_filename = fill_excel(template_source_path, mapping, selected_template['mapping'], output_name=custom_filename)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

    # Cleanup input files in background
    background_tasks.add_task(cleanup_files, file_paths)

    return {"filename": output_filename}

@app.get("/download/{filename}")
async def download_file(filename: str, background_tasks: BackgroundTasks, username: str = Depends(get_current_username)):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    # Schedule cleanup of output file after response
    background_tasks.add_task(cleanup_files, [file_path])
    
    return FileResponse(file_path, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == "__main__":
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
